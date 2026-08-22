"""Download seletivo de arquivos a partir do inventário Excel do DriveAutomate."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import time
import unicodedata
import uuid
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterator
from urllib.parse import quote

from drive_inventory import (
    FOLDER_MIME_TYPE,
    SHORTCUT_MIME_TYPE,
    WORKSPACE_PREFIX,
    build_drive_service,
    execute_drive_request,
    human_size,
)

DOWNLOAD_FIELDS = (
    "id,name,mimeType,size,modifiedTime,version,headRevisionId,resourceKey,"
    "md5Checksum,sha256Checksum,trashed,capabilities(canDownload),"
    "shortcutDetails(targetId,targetMimeType,targetResourceKey)"
)
DOWNLOAD_CHUNK_SIZE = 4 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 8 * 1024 * 1024 * 1024
MAX_XLSX_MEMBERS = 20_000
MAX_COMPONENT_LENGTH = 120
SAFE_PATH_BUDGET = 240
MIN_FREE_SPACE_MARGIN = 64 * 1024 * 1024

TRUE_MARKERS = {"sim", "s", "yes", "y", "1", "true", "x", "baixar"}
INVALID_FILENAME_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
DRIVE_IDENTIFIER_RE = re.compile(r"^[A-Za-z0-9_-]+$")
WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}

EXPORT_FORMATS: dict[str, dict[str, tuple[str, str]]] = {
    "application/vnd.google-apps.document": {
        "AUTO": (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".docx",
        ),
        "DOCX": (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".docx",
        ),
        "PDF": ("application/pdf", ".pdf"),
    },
    "application/vnd.google-apps.spreadsheet": {
        "AUTO": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xlsx",
        ),
        "XLSX": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xlsx",
        ),
        "PDF": ("application/pdf", ".pdf"),
    },
    "application/vnd.google-apps.presentation": {
        "AUTO": (
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            ".pptx",
        ),
        "PPTX": (
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            ".pptx",
        ),
        "PDF": ("application/pdf", ".pdf"),
    },
    "application/vnd.google-apps.drawing": {
        "AUTO": ("image/png", ".png"),
        "PNG": ("image/png", ".png"),
        "PDF": ("application/pdf", ".pdf"),
    },
    "application/vnd.google-apps.form": {
        "AUTO": ("application/zip", ".zip"),
        "ZIP": ("application/zip", ".zip"),
    },
    "application/vnd.google-apps.script": {
        "AUTO": ("application/vnd.google-apps.script+json", ".json"),
        "JSON": ("application/vnd.google-apps.script+json", ".json"),
    },
    "application/vnd.google-apps.site": {
        "AUTO": ("text/raw", ".txt"),
        "TXT": ("text/raw", ".txt"),
    },
    "application/vnd.google-apps.vid": {
        "AUTO": ("video/mp4", ".mp4"),
        "MP4": ("video/mp4", ".mp4"),
    },
    "application/vnd.google-apps.jam": {
        "AUTO": ("application/pdf", ".pdf"),
        "PDF": ("application/pdf", ".pdf"),
    },
}
LRO_NATIVE_MIME_TYPES = {
    "application/vnd.google-apps.form",
    "application/vnd.google-apps.script",
    "application/vnd.google-apps.site",
    "application/vnd.google-apps.vid",
    "application/vnd.google-apps.jam",
}


class DownloadCancelled(RuntimeError):
    """Interrompe o download cooperativamente."""


class InventoryWorkbookError(ValueError):
    """Indica que a planilha não segue um formato seguro/suportado."""


@dataclass(frozen=True)
class DownloadEntry:
    sheet_name: str
    row_number: int
    item_id: str
    name: str
    mime_type: str
    item_type: str
    size_bytes: int | None
    path_segments: tuple[str, ...]
    export_format: str = "AUTO"
    resource_key: str = ""
    shortcut_target_id: str = ""
    shortcut_target_mime_type: str = ""
    shortcut_target_resource_key: str = ""


@dataclass
class WorkbookScan:
    matching_sheets: int = 0
    total_rows: int = 0
    candidates: int = 0
    folders_ignored: int = 0
    selected_bytes_known: int = 0
    invalid_rows: int = 0

    def message(self) -> str:
        size = human_size(self.selected_bytes_known)
        return (
            f"{self.candidates:,} arquivo(s) selecionado(s), "
            f"{self.folders_ignored:,} pasta(s) ignorada(s), "
            f"{self.invalid_rows:,} linha(s) inválida(s). "
            f"Tamanho conhecido: {size or 'não informado'}."
        )


@dataclass
class DownloadStats:
    requested: int = 0
    downloaded: int = 0
    skipped_existing: int = 0
    renamed_existing: int = 0
    skipped_unsupported: int = 0
    failed: int = 0
    bytes_written: int = 0
    report_path: Path | None = None


def normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char for char in text.casefold() if char.isalnum())


def is_selected(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in TRUE_MARKERS


def _validate_xlsx_container(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        raise InventoryWorkbookError("Selecione um arquivo Excel no formato .xlsx.")
    if not path.is_file() or not zipfile.is_zipfile(path):
        raise InventoryWorkbookError("O arquivo selecionado não é uma planilha .xlsx válida.")

    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise InventoryWorkbookError("A planilha contém arquivos internos demais.")
            total_size = sum(member.file_size for member in members)
            if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise InventoryWorkbookError("A planilha é grande demais para ser processada com segurança.")
            compressed = sum(max(member.compress_size, 1) for member in members)
            if total_size > 100 * 1024 * 1024 and total_size / compressed > 1_000:
                raise InventoryWorkbookError("A compactação da planilha parece inválida ou maliciosa.")
    except zipfile.BadZipFile as exc:
        raise InventoryWorkbookError("A planilha está corrompida.") from exc


def _cell_text(value: object, field_name: str, sheet: str, row: int) -> str:
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("="):
        raise InventoryWorkbookError(
            f"A célula '{field_name}' em {sheet}!{row} contém fórmula. "
            "IDs e caminhos precisam ser valores fixos."
        )
    return str(value).strip()


def _drive_identifier(value: object, field_name: str, sheet: str, row: int) -> str:
    text = _cell_text(value, field_name, sheet, row)
    if text and not DRIVE_IDENTIFIER_RE.fullmatch(text):
        raise InventoryWorkbookError(
            f"A célula '{field_name}' em {sheet}!{row} não contém um identificador válido."
        )
    return text


def _parse_size(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        result = int(value)
    except (TypeError, ValueError):
        return None
    return result if result >= 0 else None


def _parse_path_segments(raw_json: object, display_path: object, name: str) -> tuple[str, ...]:
    if isinstance(raw_json, str) and raw_json and not raw_json.startswith("="):
        try:
            values = json.loads(raw_json)
        except json.JSONDecodeError:
            values = None
        if (
            isinstance(values, list)
            and values
            and all(isinstance(value, str) for value in values)
        ):
            return tuple(values)

    display = str(display_path or "").strip()
    segments = tuple(segment for segment in display.replace("\\", "/").split("/") if segment)
    if segments:
        return segments
    return (name,)


def _sheet_header_map(first_row: tuple[object, ...]) -> dict[str, int]:
    return {
        normalize_header(value): index
        for index, value in enumerate(first_row)
        if normalize_header(value)
    }


def iter_download_entries(
    workbook_path: str | Path,
    *,
    selected_only: bool = True,
    scan: WorkbookScan | None = None,
) -> Iterator[DownloadEntry]:
    """Lê abas por cabeçalho e nunca avalia fórmulas do Excel."""
    from openpyxl import load_workbook

    path = Path(workbook_path)
    _validate_xlsx_container(path)
    result = scan if scan is not None else WorkbookScan()
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    required = {"id", "mimetype", "nome", "caminhocompleto"}
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            first_row = next(rows, None)
            if not first_row:
                continue
            headers = _sheet_header_map(first_row)
            if not required.issubset(headers):
                continue
            if selected_only and "baixar" not in headers:
                raise InventoryWorkbookError(
                    "A planilha não possui a coluna 'Baixar?'. Use o modo "
                    "'Todos os arquivos restantes' para relatórios antigos."
                )

            result.matching_sheets += 1
            for row_number, values in enumerate(rows, start=2):
                result.total_rows += 1
                if selected_only:
                    marker = values[headers["baixar"]] if headers["baixar"] < len(values) else None
                    if not is_selected(marker):
                        continue

                try:
                    item_id = _drive_identifier(
                        values[headers["id"]], "ID", worksheet.title, row_number
                    )
                    name = _cell_text(values[headers["nome"]], "Nome", worksheet.title, row_number)
                    mime_type = _cell_text(
                        values[headers["mimetype"]], "MIME type", worksheet.title, row_number
                    )
                    display_path = values[headers["caminhocompleto"]]
                except (IndexError, InventoryWorkbookError):
                    result.invalid_rows += 1
                    continue
                if not item_id or not name or not mime_type:
                    result.invalid_rows += 1
                    continue

                item_type = ""
                if "tipo" in headers and headers["tipo"] < len(values):
                    item_type = str(values[headers["tipo"]] or "").strip()
                if mime_type == FOLDER_MIME_TYPE or item_type.casefold() == "pasta":
                    result.folders_ignored += 1
                    continue

                def optional(header: str) -> object:
                    index = headers.get(header)
                    return values[index] if index is not None and index < len(values) else ""

                path_segments = _parse_path_segments(
                    optional("caminhojson"), display_path, name
                )
                export_format = str(optional("formatodeexportacao") or "AUTO").strip().upper()
                size_bytes = _parse_size(optional("tamanhobytes"))
                result.candidates += 1
                result.selected_bytes_known += size_bytes or 0
                yield DownloadEntry(
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    item_id=item_id,
                    name=name,
                    mime_type=mime_type,
                    item_type=item_type,
                    size_bytes=size_bytes,
                    path_segments=path_segments,
                    export_format=export_format or "AUTO",
                    resource_key=_drive_identifier(
                        optional("resourcekey"), "Resource key", worksheet.title, row_number
                    ),
                    shortcut_target_id=_drive_identifier(
                        optional("atalhoiddoalvo"),
                        "Atalho: ID do alvo",
                        worksheet.title,
                        row_number,
                    ),
                    shortcut_target_mime_type=str(
                        optional("atalhomimetypedoalvo") or ""
                    ).strip(),
                    shortcut_target_resource_key=_drive_identifier(
                        optional("atalhoresourcekeydoalvo"),
                        "Atalho: resource key do alvo",
                        worksheet.title,
                        row_number,
                    ),
                )
    finally:
        workbook.close()

    if result.matching_sheets == 0:
        raise InventoryWorkbookError(
            "Nenhuma aba de inventário válida foi encontrada. São necessárias as "
            "colunas ID, MIME type, Nome e Caminho completo."
        )


def scan_inventory_workbook(
    workbook_path: str | Path,
    *,
    selected_only: bool = True,
) -> WorkbookScan:
    scan = WorkbookScan()
    for _entry in iter_download_entries(
        workbook_path, selected_only=selected_only, scan=scan
    ):
        pass
    return scan


def sanitize_path_component(value: str, item_id: str = "") -> str:
    text = unicodedata.normalize("NFKC", value or "")
    text = INVALID_FILENAME_RE.sub("_", text).strip(" .")
    if text in {"", ".", ".."}:
        text = "sem_nome"
    stem = text.split(".", maxsplit=1)[0].upper()
    if stem in WINDOWS_RESERVED_NAMES:
        text = f"_{text}"
    if len(text) > MAX_COMPONENT_LENGTH:
        suffix = Path(text).suffix[:20]
        digest = hashlib.sha256(f"{value}|{item_id}".encode("utf-8")).hexdigest()[:8]
        budget = MAX_COMPONENT_LENGTH - len(suffix) - len(digest) - 2
        text = f"{text[:max(budget, 1)].rstrip(' .')}~{digest}{suffix}"
    return text


def _is_reparse_point(path: Path) -> bool:
    try:
        stat_result = path.stat()
    except OSError:
        return False
    attributes = getattr(stat_result, "st_file_attributes", 0)
    return bool(attributes & 0x400) or path.is_symlink()


def safe_destination_path(
    destination_root: str | Path,
    directory_segments: tuple[str, ...],
    filename: str,
    item_id: str,
    claimed_paths: set[str] | None = None,
) -> Path:
    root = Path(destination_root).resolve()
    root.mkdir(parents=True, exist_ok=True)
    segments = [sanitize_path_component(segment, item_id) for segment in directory_segments]
    safe_name = sanitize_path_component(filename, item_id)
    candidate = root.joinpath(*segments, safe_name)

    if len(str(candidate)) > SAFE_PATH_BUDGET:
        digest = hashlib.sha256(
            "/".join((*directory_segments, filename)).encode("utf-8")
        ).hexdigest()[:12]
        first = segments[0] if segments else "arquivos"
        candidate = root / first / f"_caminho_longo_{digest}" / safe_name

    resolved = candidate.resolve(strict=False)
    try:
        if os.path.commonpath([str(root), str(resolved)]) != str(root):
            raise InventoryWorkbookError("Um caminho da planilha tenta sair da pasta de destino.")
    except ValueError as exc:
        raise InventoryWorkbookError("Um caminho da planilha aponta para outro disco.") from exc

    current = root
    for segment in resolved.relative_to(root).parts[:-1]:
        current = current / segment
        if current.exists() and _is_reparse_point(current):
            raise InventoryWorkbookError(
                "O destino contém um link/junção e não é seguro para este download."
            )

    claimed = claimed_paths if claimed_paths is not None else set()
    key = unicodedata.normalize("NFKC", str(resolved)).casefold()
    collision_index = 0
    while key in claimed:
        collision_index += 1
        suffix = resolved.suffix
        label = item_id[:8] or "duplicado"
        counter = f"-{collision_index}" if collision_index > 1 else ""
        base = candidate.stem
        resolved = candidate.with_name(f"{base} [{label}{counter}]{suffix}").resolve(
            strict=False
        )
        key = unicodedata.normalize("NFKC", str(resolved)).casefold()
    claimed.add(key)
    return resolved


def _resource_key_header(item_id: str, resource_key: str) -> dict[str, str]:
    if not resource_key:
        return {}
    return {"X-Goog-Drive-Resource-Keys": f"{item_id}/{resource_key}"}


def _get_metadata(service, item_id: str, resource_key: str = "") -> dict[str, Any]:
    request = service.files().get(
        fileId=item_id,
        fields=DOWNLOAD_FIELDS,
        supportsAllDrives=True,
    )
    request.headers.update(_resource_key_header(item_id, resource_key))
    return execute_drive_request(request, item_id)


def resolve_download_metadata(service, entry: DownloadEntry) -> tuple[dict[str, Any], str]:
    metadata = _get_metadata(service, entry.item_id, entry.resource_key)
    if metadata.get("mimeType") != SHORTCUT_MIME_TYPE:
        metadata["_driveautomateResourceKey"] = (
            metadata.get("resourceKey") or entry.resource_key
        )
        return metadata, metadata.get("name") or entry.name

    shortcut = metadata.get("shortcutDetails", {})
    target_id = shortcut.get("targetId") or entry.shortcut_target_id
    target_key = shortcut.get("targetResourceKey") or entry.shortcut_target_resource_key
    if not target_id:
        raise RuntimeError("O atalho não possui um arquivo de destino acessível.")
    target = _get_metadata(service, target_id, target_key)
    target["_driveautomateResourceKey"] = target.get("resourceKey") or target_key
    return target, entry.name or target.get("name", "arquivo")


def _choose_export(
    source_mime: str,
    requested_format: str,
    available_formats: dict[str, list[str]],
) -> tuple[str, str]:
    choices = EXPORT_FORMATS.get(source_mime)
    requested = (requested_format or "AUTO").upper()
    if not choices or requested not in choices:
        raise RuntimeError(
            f"O tipo Google Workspace '{source_mime}' não suporta o formato {requested}."
        )
    export_mime, extension = choices[requested]
    supported = available_formats.get(source_mime)
    if supported is not None and export_mime not in supported:
        if source_mime in LRO_NATIVE_MIME_TYPES:
            return export_mime, extension
        raise RuntimeError(
            f"A conta Google não oferece exportação de '{source_mime}' para {requested}."
        )
    return export_mime, extension


def _with_extension(filename: str, extension: str) -> str:
    return filename if filename.casefold().endswith(extension.casefold()) else filename + extension


def _signature(metadata: dict[str, Any]) -> dict[str, str]:
    return {
        "id": str(metadata.get("id", "")),
        "size": str(metadata.get("size", "")),
        "modifiedTime": str(metadata.get("modifiedTime", "")),
        "version": str(metadata.get("version", "")),
    }


def _write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        temporary.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _verify_checksum(path: Path, metadata: dict[str, Any]) -> None:
    expected_sha256 = metadata.get("sha256Checksum")
    expected_md5 = metadata.get("md5Checksum")
    if not expected_sha256 and not expected_md5:
        return
    sha256 = hashlib.sha256() if expected_sha256 else None
    md5 = hashlib.md5() if expected_md5 else None  # noqa: S324 - integrity, not security
    with path.open("rb") as stream:
        while chunk := stream.read(DOWNLOAD_CHUNK_SIZE):
            if sha256:
                sha256.update(chunk)
            if md5:
                md5.update(chunk)
    if sha256 and sha256.hexdigest().casefold() != str(expected_sha256).casefold():
        raise RuntimeError("A verificação SHA-256 do arquivo baixado falhou.")
    if md5 and md5.hexdigest().casefold() != str(expected_md5).casefold():
        raise RuntimeError("A verificação MD5 do arquivo baixado falhou.")


def _ensure_free_space(directory: Path, required_bytes: int | None) -> None:
    if required_bytes is None or required_bytes <= 0:
        return
    try:
        free = shutil.disk_usage(directory).free
    except OSError:
        return
    required_with_margin = required_bytes + MIN_FREE_SPACE_MARGIN
    if free < required_with_margin:
        raise RuntimeError(
            "Espaço insuficiente no disco de destino. "
            f"Necessário agora: {human_size(required_with_margin)}; "
            f"livre: {human_size(free)}."
        )


def _authorized_session(credentials):
    from google.auth.transport.requests import AuthorizedSession
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry

    session = AuthorizedSession(credentials)
    retries = Retry(
        total=5,
        connect=5,
        read=5,
        status=5,
        allowed_methods=frozenset({"GET"}),
        status_forcelist=(429, 500, 502, 503, 504),
        backoff_factor=1,
        respect_retry_after_header=True,
        raise_on_status=False,
    )
    session.mount("https://", HTTPAdapter(max_retries=retries))
    return session


def _validate_resumed_response(response, offset: int, expected_size: int | None) -> None:
    if not offset or response.status_code != 206:
        return
    content_range = response.headers.get("Content-Range", "")
    match = re.fullmatch(r"bytes (\d+)-(\d+)/(\d+|\*)", content_range)
    if not match or int(match.group(1)) != offset:
        raise RuntimeError("O servidor retornou uma faixa inválida; o parcial foi descartado.")
    remote_size = match.group(3)
    if expected_size is not None and remote_size != "*" and int(remote_size) != expected_size:
        raise RuntimeError("O arquivo mudou no Google Drive durante a retomada.")


def download_blob(
    credentials,
    metadata: dict[str, Any],
    destination: Path,
    *,
    resource_key: str = "",
    cancel_callback: Callable[[], bool] | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
    latest_metadata_callback: Callable[[], dict[str, Any]] | None = None,
) -> int:
    """Baixa blob com Range e retoma um parcial compatível após reinício."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    part = destination.with_name(f".{destination.name}.driveautomate.part")
    state = part.with_suffix(part.suffix + ".json")
    signature = _signature(metadata)
    expected_size = int(metadata["size"]) if metadata.get("size") not in (None, "") else None

    if part.exists():
        try:
            saved_signature = json.loads(state.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            saved_signature = None
        if saved_signature != signature or (expected_size is not None and part.stat().st_size > expected_size):
            part.unlink(missing_ok=True)
            state.unlink(missing_ok=True)
    _write_json_atomic(state, signature)

    offset = part.stat().st_size if part.exists() else 0
    _ensure_free_space(
        destination.parent,
        max(expected_size - offset, 0) if expected_size is not None else None,
    )
    item_id = str(metadata["id"])
    url = f"https://www.googleapis.com/drive/v3/files/{quote(item_id)}"
    params = {"alt": "media", "supportsAllDrives": "true"}
    headers = _resource_key_header(item_id, resource_key)
    if offset:
        headers["Range"] = f"bytes={offset}-"

    if cancel_callback and cancel_callback():
        raise DownloadCancelled("Download cancelado pelo usuário.")

    session = _authorized_session(credentials)
    try:
        response = session.get(
            url,
            params=params,
            headers=headers,
            stream=True,
            timeout=(20, 120),
        )
        if offset and response.status_code == 200:
            offset = 0
            part.unlink(missing_ok=True)
            _ensure_free_space(destination.parent, expected_size)
        elif response.status_code == 416 and expected_size is not None and offset == expected_size:
            response.close()
            try:
                _verify_checksum(part, metadata)
                if latest_metadata_callback and _signature(latest_metadata_callback()) != signature:
                    raise RuntimeError("O arquivo mudou no Google Drive durante o download.")
            except Exception:
                part.unlink(missing_ok=True)
                state.unlink(missing_ok=True)
                raise
            os.replace(part, destination)
            state.unlink(missing_ok=True)
            return expected_size
        response.raise_for_status()

        try:
            _validate_resumed_response(response, offset, expected_size)
        except Exception:
            response.close()
            part.unlink(missing_ok=True)
            state.unlink(missing_ok=True)
            raise

        mode = "ab" if offset and response.status_code == 206 else "wb"
        written = offset if mode == "ab" else 0
        with part.open(mode) as stream:
            for chunk in response.iter_content(chunk_size=DOWNLOAD_CHUNK_SIZE):
                if cancel_callback and cancel_callback():
                    raise DownloadCancelled("Download cancelado. O arquivo parcial foi preservado.")
                if not chunk:
                    continue
                stream.write(chunk)
                written += len(chunk)
                if progress_callback:
                    progress_callback(written, expected_size or 0)
    finally:
        session.close()

    if expected_size is not None and part.stat().st_size != expected_size:
        raise RuntimeError(
            f"Download incompleto: esperado {expected_size} bytes, recebido {part.stat().st_size}."
        )
    try:
        _verify_checksum(part, metadata)
        if latest_metadata_callback and _signature(latest_metadata_callback()) != signature:
            raise RuntimeError("O arquivo mudou no Google Drive durante o download.")
    except Exception:
        part.unlink(missing_ok=True)
        state.unlink(missing_ok=True)
        raise
    os.replace(part, destination)
    state.unlink(missing_ok=True)
    return destination.stat().st_size


def _wait_for_operation(
    service,
    operation: dict[str, Any],
    *,
    item_id: str,
    resource_key: str,
    cancel_callback: Callable[[], bool] | None,
) -> dict[str, Any]:
    operation_name = str(operation.get("name") or "")
    delay = 1.0
    while not operation.get("done"):
        if cancel_callback and cancel_callback():
            raise DownloadCancelled("Download cancelado pelo usuário.")
        if not operation_name:
            raise RuntimeError("O Google não retornou o identificador da preparação do arquivo.")
        deadline = time.monotonic() + delay
        while time.monotonic() < deadline:
            if cancel_callback and cancel_callback():
                raise DownloadCancelled("Download cancelado pelo usuário.")
            time.sleep(min(0.25, max(deadline - time.monotonic(), 0)))
        request = service.operations().get(name=operation_name)
        request.headers.update(_resource_key_header(item_id, resource_key))
        operation = execute_drive_request(request, item_id)
        delay = min(delay * 2, 10.0)
    return operation


def download_workspace_file(
    service,
    credentials,
    metadata: dict[str, Any],
    destination: Path,
    export_mime: str,
    *,
    resource_key: str = "",
    cancel_callback: Callable[[], bool] | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
) -> int:
    """Prepara e baixa conteúdo Workspace pela operação longa files.download."""
    item_id = str(metadata["id"])
    request = service.files().download(fileId=item_id, mimeType=export_mime)
    request.headers.update(_resource_key_header(item_id, resource_key))
    operation = execute_drive_request(request, item_id)
    operation = _wait_for_operation(
        service,
        operation,
        item_id=item_id,
        resource_key=resource_key,
        cancel_callback=cancel_callback,
    )
    if operation.get("error"):
        error = operation["error"]
        message = str(error.get("message") or "O Google não conseguiu preparar o arquivo.")
        raise RuntimeError(message)
    response_data = operation.get("response") or {}
    download_uri = str(response_data.get("downloadUri") or "")
    if not download_uri:
        raise RuntimeError("O Google concluiu a preparação sem fornecer o link de download.")

    destination.parent.mkdir(parents=True, exist_ok=True)
    part = destination.with_name(f".{destination.name}.driveautomate-export.part")
    part.unlink(missing_ok=True)
    session = _authorized_session(credentials)
    try:
        if cancel_callback and cancel_callback():
            raise DownloadCancelled("Download cancelado pelo usuário.")
        response = session.get(
            download_uri,
            headers=_resource_key_header(item_id, resource_key),
            stream=True,
            timeout=(20, 120),
            allow_redirects=True,
        )
        response.raise_for_status()
        try:
            total = int(response.headers.get("Content-Length", ""))
        except (TypeError, ValueError):
            total = 0
        _ensure_free_space(destination.parent, total or None)
        written = 0
        with part.open("wb") as stream:
            for chunk in response.iter_content(chunk_size=DOWNLOAD_CHUNK_SIZE):
                if cancel_callback and cancel_callback():
                    raise DownloadCancelled("Download cancelado pelo usuário.")
                if not chunk:
                    continue
                stream.write(chunk)
                written += len(chunk)
                if progress_callback:
                    progress_callback(written, total)
        if total and written != total:
            raise RuntimeError(
                f"Download incompleto: esperado {total} bytes, recebido {written}."
            )
        os.replace(part, destination)
        return written
    except Exception:
        part.unlink(missing_ok=True)
        raise
    finally:
        session.close()


def export_workspace_file(
    service,
    metadata: dict[str, Any],
    destination: Path,
    export_mime: str,
    *,
    resource_key: str = "",
    cancel_callback: Callable[[], bool] | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
) -> int:
    """Fallback legado: files.export possui limite de 10 MB por arquivo."""
    from googleapiclient.http import MediaIoBaseDownload

    destination.parent.mkdir(parents=True, exist_ok=True)
    part = destination.with_name(f".{destination.name}.driveautomate-export.part")
    part.unlink(missing_ok=True)
    request = service.files().export_media(
        fileId=metadata["id"], mimeType=export_mime
    )
    request.headers.update(_resource_key_header(metadata["id"], resource_key))
    try:
        with part.open("wb") as stream:
            downloader = MediaIoBaseDownload(stream, request, chunksize=DOWNLOAD_CHUNK_SIZE)
            done = False
            while not done:
                if cancel_callback and cancel_callback():
                    raise DownloadCancelled("Download cancelado pelo usuário.")
                status, done = downloader.next_chunk(num_retries=5)
                if progress_callback and status:
                    progress_callback(int(status.resumable_progress), int(status.total_size or 0))
        os.replace(part, destination)
        return destination.stat().st_size
    except Exception:
        part.unlink(missing_ok=True)
        raise


def _safe_error(exc: BaseException) -> str:
    text = " ".join(str(exc).split())
    return text[:1_000] or exc.__class__.__name__


def _csv_safe(value: object) -> str:
    text = str(value if value is not None else "")
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _write_report_row(writer, values) -> None:
    writer.writerow([_csv_safe(value) for value in values])


@contextmanager
def _download_report(destination: Path):
    final_path = destination / "driveautomate_downloads.csv"
    partial_path = destination / "driveautomate_downloads_parcial.csv"
    temporary = destination / f".driveautomate_downloads.{uuid.uuid4().hex}.tmp"
    try:
        with temporary.open("w", encoding="utf-8-sig", newline="") as report:
            yield csv.writer(report), final_path
    except DownloadCancelled:
        if temporary.exists():
            os.replace(temporary, partial_path)
        raise
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    else:
        os.replace(temporary, final_path)


def download_from_inventory(
    workbook_path: str | Path,
    destination_root: str | Path,
    credentials,
    *,
    selected_only: bool = True,
    skip_existing: bool = True,
    log_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[int, int, int], None] | None = None,
    cancel_callback: Callable[[], bool] | None = None,
) -> DownloadStats:
    """Valida a planilha, reconsulta metadados e baixa cada item selecionado."""
    scan = scan_inventory_workbook(workbook_path, selected_only=selected_only)
    if scan.candidates == 0:
        raise InventoryWorkbookError(
            "Nenhum arquivo foi selecionado. Marque 'Baixar?' como Sim ou escolha "
            "o modo de baixar todas as linhas restantes."
        )

    service = build_drive_service(credentials)
    about = execute_drive_request(service.about().get(fields="exportFormats"), "exportFormats")
    available_formats = about.get("exportFormats", {})
    destination = Path(destination_root).resolve()
    destination.mkdir(parents=True, exist_ok=True)
    claimed_paths: set[str] = set()
    stats = DownloadStats(requested=scan.candidates)

    with _download_report(destination) as (writer, report_path):
        _write_report_row(writer, ["Status", "Arquivo", "ID", "Destino", "Detalhes"])
        for current, entry in enumerate(
            iter_download_entries(workbook_path, selected_only=selected_only), start=1
        ):
            if cancel_callback and cancel_callback():
                raise DownloadCancelled("Download cancelado pelo usuário.")
            if log_callback:
                log_callback(f"Preparando {current:,}/{scan.candidates:,}: {entry.name}")
            try:
                metadata, local_name = resolve_download_metadata(service, entry)
                if metadata.get("trashed"):
                    raise RuntimeError("O arquivo está na lixeira do Google Drive.")
                if metadata.get("mimeType") == FOLDER_MIME_TYPE:
                    raise RuntimeError("O atalho aponta para uma pasta; pastas não são baixadas.")
                if metadata.get("capabilities", {}).get("canDownload") is False:
                    raise RuntimeError("O proprietário bloqueou o download deste arquivo.")

                source_mime = metadata.get("mimeType", "")
                export_mime = ""
                resource_key = str(metadata.get("_driveautomateResourceKey") or "")
                if source_mime.startswith(WORKSPACE_PREFIX):
                    export_mime, extension = _choose_export(
                        source_mime, entry.export_format, available_formats
                    )
                    local_name = _with_extension(local_name, extension)

                directory_segments = entry.path_segments[:-1]
                output = safe_destination_path(
                    destination,
                    directory_segments,
                    local_name,
                    str(metadata.get("id") or entry.item_id),
                    claimed_paths,
                )
                if output.exists() and skip_existing:
                    stats.skipped_existing += 1
                    _write_report_row(
                        writer,
                        ["Ignorado", local_name, entry.item_id, str(output), "Já existe"],
                    )
                    continue
                if output.exists():
                    while output.exists():
                        output = safe_destination_path(
                            destination,
                            directory_segments,
                            local_name,
                            str(metadata.get("id") or entry.item_id),
                            claimed_paths,
                        )
                    stats.renamed_existing += 1
                    if log_callback:
                        log_callback(f"Conflito: será salvo como {output.name}")

                def file_progress(written: int, total: int) -> None:
                    percentage = int((written / total) * 100) if total else 0
                    if progress_callback:
                        progress_callback(current, scan.candidates, percentage)

                if export_mime:
                    files_resource = service.files()
                    if callable(getattr(files_resource, "download", None)):
                        written = download_workspace_file(
                            service,
                            credentials,
                            metadata,
                            output,
                            export_mime,
                            resource_key=resource_key,
                            cancel_callback=cancel_callback,
                            progress_callback=file_progress,
                        )
                    else:  # compatibilidade com clientes antigos da API
                        written = export_workspace_file(
                            service,
                            metadata,
                            output,
                            export_mime,
                            resource_key=resource_key,
                            cancel_callback=cancel_callback,
                            progress_callback=file_progress,
                        )
                else:
                    written = download_blob(
                        credentials,
                        metadata,
                        output,
                        resource_key=resource_key,
                        cancel_callback=cancel_callback,
                        progress_callback=file_progress,
                        latest_metadata_callback=lambda: _get_metadata(
                            service, str(metadata["id"]), resource_key
                        ),
                    )
                stats.downloaded += 1
                stats.bytes_written += written
                _write_report_row(
                    writer,
                    ["Concluído", local_name, entry.item_id, str(output), f"{written} bytes"],
                )
            except DownloadCancelled:
                raise
            except RuntimeError as exc:
                message = _safe_error(exc)
                if "não suporta o formato" in message or "não oferece exportação" in message:
                    stats.skipped_unsupported += 1
                    status = "Não suportado"
                else:
                    stats.failed += 1
                    status = "Falha"
                _write_report_row(writer, [status, entry.name, entry.item_id, "", message])
                if log_callback:
                    log_callback(f"{status}: {entry.name} — {message}")
            except Exception as exc:
                stats.failed += 1
                message = _safe_error(exc)
                _write_report_row(
                    writer, ["Falha", entry.name, entry.item_id, "", message]
                )
                if log_callback:
                    log_callback(f"Falha: {entry.name} — {message}")
            finally:
                if progress_callback:
                    progress_callback(current, scan.candidates, 100)

    stats.report_path = report_path
    return stats
