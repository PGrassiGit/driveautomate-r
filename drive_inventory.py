#!/usr/bin/env python3
"""Gera um inventário hierárquico do Google Drive em Excel."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import importlib.util
import json
import os
import re
import time
import uuid
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator
from urllib.parse import parse_qs, urlparse

from oauth_config import oauth_client_path
from secure_store import is_protected_file, read_secure_json, write_secure_json

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
FOLDER_MIME_TYPE = "application/vnd.google-apps.folder"
SHORTCUT_MIME_TYPE = "application/vnd.google-apps.shortcut"
WORKSPACE_PREFIX = "application/vnd.google-apps."
REQUIRED_PACKAGES = {
    "google.auth": "google-auth",
    "google_auth_oauthlib": "google-auth-oauthlib",
    "googleapiclient": "google-api-python-client",
    "openpyxl": "openpyxl",
}
DEFAULT_FIELDS = (
    "nextPageToken,incompleteSearch,files("
    "id,name,mimeType,size,webViewLink,createdTime,modifiedTime,parents,"
    "owners(displayName,emailAddress),driveId,resourceKey,"
    "shortcutDetails(targetId,targetMimeType,targetResourceKey),"
    "capabilities(canDownload),md5Checksum,sha256Checksum,headRevisionId,version,"
    "ownedByMe,shared,trashed)"
)
MAX_EXCEL_ROWS = 1_048_576
EXCEL_HEADER_ROWS = 1
DEFAULT_PROGRESS_INTERVAL = 1_000
EXCEL_ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


class ExportCancelled(RuntimeError):
    """Interrompe uma exportação solicitada pelo usuário."""


@dataclass(frozen=True)
class DriveSource:
    """Contexto da pasta raiz e do corpus correto para pesquisá-la."""

    root_id: str
    root_name: str
    source_type: str
    drive_id: str
    drive_name: str
    owner: str
    root_resource_key: str = ""


@dataclass
class InventoryStats:
    """Contadores produzidos durante a gravação em streaming."""

    item_count: int = 0
    folder_count: int = 0
    file_count: int = 0
    total_size_bytes: int = 0
    worksheet_count: int = 0


@dataclass(frozen=True)
class DriveItem:
    """Representa uma linha do inventário."""

    root_folder: str
    source_type: str
    drive_name: str
    parent_folder: str
    path: str
    depth: int
    item_type: str
    name: str
    item_id: str
    mime_type: str
    size_bytes: int | None
    size_readable: str
    web_link: str
    created_time: str
    modified_time: str
    owners: str
    drive_id: str
    parent_id: str
    shortcut_target_id: str
    shortcut_target_mime_type: str
    path_segments_json: str = ""
    resource_key: str = ""
    shortcut_target_resource_key: str = ""
    can_download: bool | None = None
    md5_checksum: str = ""
    sha256_checksum: str = ""
    head_revision_id: str = ""
    version: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Lista pastas e arquivos do Google Drive e exporta para Excel."
    )
    parser.add_argument(
        "--credentials",
        default=str(oauth_client_path()),
        help="Caminho do JSON OAuth desktop. Padrão: configuração local do aplicativo.",
    )
    parser.add_argument(
        "--token",
        default=str(default_token_path()),
        help="Caminho onde o token OAuth será salvo. Padrão: pasta de dados do usuário.",
    )
    parser.add_argument(
        "--folder-id",
        default="root",
        help="ID ou link da pasta inicial. Use o padrão 'root' para listar o Meu Drive.",
    )
    parser.add_argument(
        "--output",
        default="relatorio_google_drive.xlsx",
        help="Nome do arquivo Excel de saída. Padrão: relatorio_google_drive.xlsx",
    )
    parser.add_argument(
        "--include-shared-drives",
        action="store_true",
        default=True,
        help="Mantém o suporte automático a itens e Drives compartilhados (padrão).",
    )
    parser.add_argument(
        "--personal-only",
        action="store_false",
        dest="include_shared_drives",
        help="Desativa o suporte a Drives compartilhados.",
    )
    return parser.parse_args()


def default_token_path() -> Path:
    """Retorna um local persistente para token, inclusive quando rodar como .exe."""
    appdata = os.getenv("APPDATA")
    if appdata:
        return Path(appdata) / "DriveAutomate" / "token.dat"

    return Path.home() / ".driveautomate" / "token.dat"


def legacy_default_token_path() -> Path:
    appdata = os.getenv("APPDATA")
    if appdata:
        return Path(appdata) / "DriveAutomate" / "token.json"
    return Path.home() / ".driveautomate" / "token.json"


def module_available(module_name: str) -> bool:
    """Verifica módulos aninhados sem importar pacotes de terceiros."""
    parent_name = module_name.split(".", maxsplit=1)[0]
    if importlib.util.find_spec(parent_name) is None:
        return False
    return importlib.util.find_spec(module_name) is not None


def ensure_dependencies() -> None:
    """Mostra uma mensagem amigável quando as dependências não foram instaladas."""
    missing = [
        package_name
        for module_name, package_name in REQUIRED_PACKAGES.items()
        if not module_available(module_name)
    ]
    if missing:
        packages = ", ".join(sorted(missing))
        raise SystemExit(
            "Dependências Python não instaladas: "
            f"{packages}. Execute: python -m pip install -r requirements.txt"
        )


def default_accounts_dir() -> Path:
    """Retorna o diretório onde cada conta autenticada guarda seu token."""
    appdata = os.getenv("APPDATA")
    if appdata:
        return Path(appdata) / "DriveAutomate" / "accounts"

    return Path.home() / ".driveautomate" / "accounts"


def sanitize_account_filename(value: str) -> str:
    """Gera um identificador estável sem expor o e-mail no nome do arquivo."""
    normalized = value.strip().casefold().encode("utf-8")
    return hashlib.sha256(normalized).hexdigest()[:32]


def account_token_path(email: str) -> Path:
    """Retorna o caminho do token para um e-mail autenticado."""
    return default_accounts_dir() / f"{sanitize_account_filename(email)}.dat"


def token_identity(token_path: str | Path) -> tuple[str, str]:
    """Lê somente a identificação local associada ao token protegido."""
    try:
        payload = read_secure_json(token_path)
    except (OSError, ValueError):
        return "", ""
    return (
        str(payload.get("account_email") or ""),
        str(payload.get("account_display_name") or ""),
    )


def format_drive_error(exc: BaseException, folder_id: str) -> str:
    """Formata erros de API do Drive para uma mensagem amigável ao usuário."""
    error_text = " ".join(
        str(arg).lower() for arg in (exc, *getattr(exc, "args", ())) if str(arg)
    )
    status = getattr(getattr(exc, "resp", None), "status", None)

    if status in {429} or "ratelimit" in error_text or "rate limit" in error_text:
        return (
            "O Google Drive limitou temporariamente a velocidade das consultas. "
            "Aguarde alguns minutos e tente novamente."
        )
    if status and 500 <= status <= 599:
        return (
            "O Google Drive está temporariamente indisponível. "
            "Aguarde alguns minutos e tente novamente."
        )
    if status in {403} or "insufficientpermissions" in error_text or "permission denied" in error_text:
        return (
            "Não foi possível acessar esta pasta. Verifique se o link ou ID pertence "
            "à conta Google autenticada e se essa conta tem permissão para ver o conteúdo. "
            "Se precisar, use Gerenciar contas para trocar de conta."
        )
    if status in {404} or "notfound" in error_text or "file not found" in error_text:
        return (
            "O link informado não foi encontrado ou não está associado à conta atual. "
            "Confira o link e tente novamente com a conta correta."
        )

    return "Não foi possível ler a pasta do Google Drive. Confira o link, a conta usada e a conexão com a internet."


def collect_error_messages(exc: BaseException) -> list[str]:
    """Coleta mensagens de exceções encadeadas para melhorar os alertas ao usuário."""
    messages: list[str] = []
    seen: set[int] = set()
    current: BaseException | None = exc

    while current is not None and id(current) not in seen:
        text = str(current).strip()
        if text:
            messages.append(text)
        seen.add(id(current))
        current = current.__cause__ or current.__context__

    return messages


def format_user_error(exc: BaseException) -> str:
    """Converte exceções técnicas em mensagens simples para o usuário final."""
    error_text = " ".join(message.lower() for message in collect_error_messages(exc))

    if isinstance(exc, ModuleNotFoundError) or "no module named" in error_text:
        return (
            "Não foi possível continuar porque as dependências do Google não estão instaladas. "
            "Instale as dependências com 'python -m pip install -r requirements.txt' e tente novamente."
        )
    if isinstance(exc, FileNotFoundError):
        return (
            "Não foi encontrado o arquivo de autenticação. "
            "Verifique se o programa foi configurado corretamente e tente novamente."
        )
    if isinstance(exc, PermissionError):
        return (
            "Não foi possível gravar no destino escolhido. Feche o Excel ou o arquivo "
            "em uso e escolha uma pasta em que você possa salvar arquivos."
        )
    if isinstance(exc, OSError) and (
        "no space" in error_text or "disk full" in error_text
    ):
        return (
            "Não há espaço livre suficiente para salvar o relatório. "
            "Libere espaço ou escolha outro disco."
        )
    if is_oauth_client_deleted_error(exc):
        return format_oauth_error(exc)
    if "browser" in error_text or "localhost" in error_text or "redirect" in error_text or "callback" in error_text:
        return (
            "A janela de login do Google não foi concluída. Tente novamente e aceite a autorização quando o navegador abrir. "
            "Se o navegador não abrir, feche e abra o programa novamente."
        )
    if "timeout" in error_text or "timed out" in error_text or "temporarily unavailable" in error_text:
        return "A autenticação demorou demais. Verifique sua internet e tente novamente."
    if "invalid_grant" in error_text or "invalid_client" in error_text or "access_denied" in error_text:
        return "A autorização foi negada pelo Google. Use a conta correta e aceite as permissões pedidas."
    if "insufficientpermissions" in error_text or "permission denied" in error_text:
        return (
            "A conta autenticada não tem permissão para abrir esta pasta. "
            "Tente trocar de conta ou pedir acesso ao compartilhamento."
        )
    if "notfound" in error_text or "file not found" in error_text:
        return (
            "A pasta ou link informado não foi encontrado. "
            "Confira o link e tente novamente."
        )
    if isinstance(exc, RuntimeError) and str(exc):
        return str(exc)

    return (
        "Não foi possível concluir a operação. Verifique a conexão, a conta Google e o link da pasta. "
        "Se o problema continuar, feche e abra o programa novamente."
    )


def execute_drive_request(request, folder_id: str) -> dict[str, Any]:
    try:
        try:
            return request.execute(num_retries=5)
        except TypeError as exc:
            # Dublês simples usados por integrações/testes podem não expor num_retries.
            if "num_retries" not in str(exc):
                raise
            return request.execute()
    except Exception as exc:
        status = getattr(getattr(exc, "resp", None), "status", None)
        if status in {403, 404} or "httperror" in exc.__class__.__name__.lower():
            raise RuntimeError(format_drive_error(exc, folder_id)) from exc
        raise


def get_drive_user_info(creds: Any) -> dict[str, str]:
    service = build_drive_service(creds)
    response = service.about().get(fields="user(displayName,emailAddress)").execute(
        num_retries=5
    )
    return response.get("user", {})


def build_drive_service(creds: Any, *, timeout_seconds: int = 120):
    """Cria um cliente isolado por worker com timeout de rede explícito."""
    import httplib2
    from google_auth_httplib2 import AuthorizedHttp
    from googleapiclient.discovery import build

    http = AuthorizedHttp(creds, http=httplib2.Http(timeout=timeout_seconds))
    return build("drive", "v3", http=http, cache_discovery=False)


def authenticate_new_account(
    credentials_path: str | None = None,
    client_config: dict[str, Any] | None = None,
) -> tuple[Path, str, str | None]:
    """Autentica uma nova conta e salva o token em um arquivo por conta."""
    try:
        ensure_dependencies()

        accounts_dir = default_accounts_dir()
        accounts_dir.mkdir(parents=True, exist_ok=True)

        temp_token = accounts_dir / f".tmp_auth_{uuid.uuid4().hex}.dat"
        try:
            creds = authenticate(
                credentials_path=credentials_path,
                token_path=temp_token,
                client_config=client_config,
            )

            user_info = get_drive_user_info(creds)
            email = user_info.get("emailAddress")
            display_name = user_info.get("displayName")
            if not email:
                raise RuntimeError(
                    "Autenticação concluída, mas não foi possível identificar o e-mail da conta."
                )

            payload = read_secure_json(temp_token)
            if "credentials" not in payload:
                payload = {"schema": 1, "credentials": payload}
            payload["account_email"] = email
            payload["account_display_name"] = display_name or ""
            write_secure_json(temp_token, payload)

            final_token = account_token_path(email)
            os.replace(temp_token, final_token)

            return final_token, email, display_name
        finally:
            if temp_token.exists():
                temp_token.unlink()
    except Exception as exc:
        raise RuntimeError(format_user_error(exc)) from exc


def is_oauth_client_deleted_error(exc: BaseException) -> bool:
    """Detecta erros de OAuth causados por um client ID/secret removido no Google Cloud."""
    error_text = " ".join(
        str(arg) for arg in (exc, *getattr(exc, "args", ())) if str(arg)
    ).lower()
    return "deleted_client" in error_text or "oauth client was deleted" in error_text


def format_oauth_error(exc: BaseException, credentials_path: str | None = None) -> str:
    """Cria uma mensagem de erro mais útil para problemas de OAuth."""
    detail = (
        "O cliente OAuth configurado não está mais válido (erro 'deleted_client'). "
        "Crie um OAuth Client ID do tipo 'Aplicativo para computador' no Google "
        "Cloud e importe o novo arquivo JSON em 'Configurar OAuth'."
    )
    if credentials_path:
        detail += (
            f" A configuração tentada foi '{credentials_path}'. Remova e conecte "
            "novamente as contas depois de trocar o cliente OAuth."
        )
    return detail


def select_oauth_credentials_source(
    credentials_path: str | None,
    client_config: dict[str, Any] | None = None,
) -> tuple[str | None, dict[str, Any] | None]:
    """Seleciona a fonte de OAuth a ser usada para autenticação."""
    if credentials_path is None and client_config is None:
        credentials_path = str(oauth_client_path())
    if credentials_path:
        credentials_file = Path(credentials_path)
        if credentials_file.exists():
            return str(credentials_file), None

    if client_config:
        return None, client_config

    if credentials_path:
        credentials_file = Path(credentials_path)
        if credentials_file.exists():
            return str(credentials_file), None

    return credentials_path, None


def authenticate(
    credentials_path: str | None,
    token_path: str | Path,
    client_config: dict[str, Any] | None = None,
) -> Any:
    """Autentica com OAuth de aplicativo desktop."""
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow

    creds = None
    token = Path(token_path)
    token_payload: dict[str, Any] = {}
    should_write = False

    if token.exists():
        try:
            token_payload = read_secure_json(token)
            nested_credentials = token_payload.get("credentials")
            token_data = (
                nested_credentials
                if isinstance(nested_credentials, dict)
                else token_payload
            )
            stored_scopes = set(token_data.get("scopes") or [])
            if set(SCOPES).issubset(stored_scopes):
                creds = Credentials.from_authorized_user_info(token_data, SCOPES)
            if not is_protected_file(token) or not isinstance(nested_credentials, dict):
                should_write = True
        except (OSError, ValueError, json.JSONDecodeError):
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                should_write = True
            except Exception as exc:  # pragma: no cover - exercised via runtime
                if is_oauth_client_deleted_error(exc):
                    raise RuntimeError(format_oauth_error(exc, credentials_path)) from exc
                raise
        else:
            credentials_file_path, resolved_client_config = select_oauth_credentials_source(
                credentials_path=credentials_path,
                client_config=client_config,
            )
            flow = None
            if credentials_file_path:
                flow = InstalledAppFlow.from_client_secrets_file(credentials_file_path, SCOPES)
            elif resolved_client_config:
                flow = InstalledAppFlow.from_client_config(resolved_client_config, SCOPES)
            else:
                raise FileNotFoundError(
                    "O cliente OAuth ainda não foi configurado. Importe um JSON "
                    "OAuth do tipo 'Aplicativo para computador' antes de conectar a conta."
                )

            try:
                creds = flow.run_local_server(port=0, timeout_seconds=300)
            except Exception as exc:  # pragma: no cover - exercised via runtime
                if is_oauth_client_deleted_error(exc):
                    raise RuntimeError(format_oauth_error(exc, credentials_path)) from exc
                raise

            should_write = True

    if should_write:
        credentials_data = json.loads(creds.to_json())
        envelope: dict[str, Any] = {"schema": 1, "credentials": credentials_data}
        for field in ("account_email", "account_display_name"):
            if token_payload.get(field):
                envelope[field] = token_payload[field]
        write_secure_json(token, envelope)

    return creds


def api_kwargs(include_shared_drives: bool) -> dict[str, object]:
    """Parâmetros de ``files.get`` para itens pessoais ou compartilhados."""
    if not include_shared_drives:
        return {}

    return {"supportsAllDrives": True}


def list_api_kwargs(
    include_shared_drives: bool,
    drive_id: str = "",
) -> dict[str, object]:
    """Parâmetros de ``files.list`` escolhidos conforme a origem da pasta.

    Pastas de um Drive compartilhado são pesquisadas apenas no corpus daquele
    Drive. Pastas do Meu Drive ou "Compartilhados comigo" usam o corpus do
    usuário. Isso evita buscas ``allDrives``, que podem ser lentas ou incompletas.
    """
    if not include_shared_drives:
        return {}

    kwargs: dict[str, object] = {
        "supportsAllDrives": True,
        "includeItemsFromAllDrives": True,
    }
    if drive_id:
        kwargs.update({"corpora": "drive", "driveId": drive_id})
    else:
        kwargs["corpora"] = "user"
    return kwargs


def resource_key_header(item_id: str, resource_key: str) -> dict[str, str]:
    if not resource_key:
        return {}
    return {"X-Goog-Drive-Resource-Keys": f"{item_id}/{resource_key}"}


def get_folder_metadata(
    service,
    folder_id: str,
    include_shared_drives: bool,
    resource_key: str = "",
) -> dict:
    fields = (
        "id,name,mimeType,webViewLink,driveId,parents,ownedByMe,shared,"
        "owners(displayName,emailAddress),resourceKey,capabilities(canListChildren)"
    )
    request = service.files().get(
        fileId=folder_id,
        fields=fields,
        **api_kwargs(include_shared_drives),
    )
    request.headers.update(resource_key_header(folder_id, resource_key))
    return execute_drive_request(request, folder_id)


def get_shared_drive_name(service, drive_id: str) -> str:
    """Obtém o nome do Drive compartilhado sem impedir a exportação em caso de falha."""
    if not drive_id:
        return ""
    try:
        response = execute_drive_request(
            service.drives().get(driveId=drive_id, fields="id,name"),
            drive_id,
        )
        return response.get("name", "")
    except Exception:
        return ""


def resolve_drive_source(
    service,
    folder_id_or_url: str,
    include_shared_drives: bool = True,
) -> DriveSource:
    """Valida a pasta e diferencia Meu Drive, compartilhamento e Shared Drive."""
    requested_id, requested_resource_key = normalize_folder_reference(folder_id_or_url)
    metadata = get_folder_metadata(
        service,
        requested_id,
        include_shared_drives,
        requested_resource_key,
    )

    if metadata.get("mimeType") != FOLDER_MIME_TYPE:
        raise RuntimeError(
            "O link informado não é de uma pasta do Google Drive. "
            "Abra a pasta no navegador, copie o link e tente novamente."
        )
    if metadata.get("capabilities", {}).get("canListChildren") is False:
        raise RuntimeError(
            "A conta autenticada consegue ver a pasta, mas não tem permissão "
            "para listar seu conteúdo."
        )

    drive_id = metadata.get("driveId", "")
    if drive_id:
        source_type = "Drive compartilhado"
        drive_name = get_shared_drive_name(service, drive_id)
    elif requested_id == "root" or metadata.get("ownedByMe") is True:
        source_type = "Meu Drive"
        drive_name = ""
    else:
        source_type = "Compartilhada comigo"
        drive_name = ""

    return DriveSource(
        root_id=metadata.get("id") or requested_id,
        root_name=metadata.get("name", "Meu Drive"),
        source_type=source_type,
        drive_id=drive_id,
        drive_name=drive_name,
        owner=owners_to_text(metadata),
        root_resource_key=metadata.get("resourceKey", "") or requested_resource_key,
    )


def normalize_folder_id(folder_id_or_url: str) -> str:
    """Aceita tanto ID puro quanto links de pasta do Google Drive."""
    return normalize_folder_reference(folder_id_or_url)[0]


def normalize_folder_reference(folder_id_or_url: str) -> tuple[str, str]:
    """Extrai ID e resource key de links normais ou compartilhados por link."""
    value = folder_id_or_url.strip()
    if not value:
        return "root", ""

    resource_key = ""
    if "://" in value:
        parsed = urlparse(value)
        query = parse_qs(parsed.query)
        resource_key = str((query.get("resourcekey") or [""])[0])
        query_id = str((query.get("id") or [""])[0])
        value = query_id or parsed.path
    elif value.startswith("id="):
        query = parse_qs(value)
        resource_key = str((query.get("resourcekey") or [""])[0])
        value = str((query.get("id") or [""])[0])

    markers = ("/folders/", "id=")
    for marker in markers:
        if marker in value:
            value = value.split(marker, maxsplit=1)[1]
            break

    for separator in ("?", "&", "/"):
        if separator in value:
            value = value.split(separator, maxsplit=1)[0]

    if resource_key and not re.fullmatch(r"[A-Za-z0-9_-]+", resource_key):
        raise RuntimeError("O link contém uma chave de compartilhamento inválida.")
    return value or "root", resource_key


def escape_query_value(value: str) -> str:
    """Escapa aspas e barras para consultas da API do Drive."""
    return value.replace("\\", "\\\\").replace("'", "\\'")


def list_children(
    service,
    folder_id: str,
    include_shared_drives: bool,
    drive_id: str = "",
    resource_key: str = "",
) -> Iterable[dict]:
    """Retorna filhos diretos de uma pasta, paginando todos os resultados."""
    page_token = None
    query = f"'{escape_query_value(folder_id)}' in parents and trashed = false"

    while True:
        request = (
            service.files()
            .list(
                q=query,
                fields=DEFAULT_FIELDS,
                pageToken=page_token,
                pageSize=1000,
                spaces="drive",
                **list_api_kwargs(include_shared_drives, drive_id),
            )
        )
        request.headers.update(resource_key_header(folder_id, resource_key))
        response = execute_drive_request(request, folder_id)

        if response.get("incompleteSearch"):
            raise RuntimeError(
                "O Google informou que a busca nesta pasta ficou incompleta. "
                "Tente novamente; o relatório não foi salvo para evitar dados faltando."
            )
        yield from response.get("files", [])

        page_token = response.get("nextPageToken")
        if not page_token:
            break


def human_size(size_bytes: int | None) -> str:
    if size_bytes is None:
        return ""

    size = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size < 1024 or unit == "TB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} B"
        size /= 1024

    return f"{size_bytes} B"


def classify_item(mime_type: str) -> str:
    if mime_type == FOLDER_MIME_TYPE:
        return "Pasta"
    if mime_type == SHORTCUT_MIME_TYPE:
        return "Atalho"
    if mime_type.startswith(WORKSPACE_PREFIX):
        return "Google Workspace"
    return "Arquivo"


def item_web_link(file_metadata: dict) -> str:
    """Retorna o link da API ou um link estável construído a partir do ID."""
    web_link = file_metadata.get("webViewLink", "")
    if web_link:
        return web_link
    item_id = file_metadata.get("id", "")
    if not item_id:
        return ""
    if file_metadata.get("mimeType") == FOLDER_MIME_TYPE:
        return f"https://drive.google.com/drive/folders/{item_id}"
    return f"https://drive.google.com/open?id={item_id}"


def parse_size(value: str | None) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def owners_to_text(file_metadata: dict) -> str:
    owners = file_metadata.get("owners", [])
    return "; ".join(
        owner.get("emailAddress") or owner.get("displayName", "") for owner in owners
    )


def iter_drive_items(
    service,
    source: DriveSource,
    include_shared_drives: bool,
    progress_callback: Callable[[str], None] | None = None,
    cancel_callback: Callable[[], bool] | None = None,
    progress_interval: int = DEFAULT_PROGRESS_INTERVAL,
) -> Iterator[DriveItem]:
    """Percorre a árvore sem guardar os arquivos encontrados em memória."""
    root_children = iter(
        list_children(
            service,
            source.root_id,
            include_shared_drives,
            drive_id=source.drive_id,
            resource_key=source.root_resource_key,
        )
    )
    folder_stack: list[
        tuple[str, str, str, tuple[str, ...], str, int, Iterator[dict]]
    ] = [
        (
            source.root_id,
            source.root_name,
            source.root_name,
            (source.root_name,),
            source.root_resource_key,
            0,
            root_children,
        )
    ]
    visited_folders: set[str] = {source.root_id}
    last_folder_progress = 0
    item_count = 0
    started = time.monotonic()

    while folder_stack:
        if cancel_callback and cancel_callback():
            raise ExportCancelled("Exportação cancelada pelo usuário.")

        (
            current_id,
            current_name,
            current_path,
            current_segments,
            _current_resource_key,
            depth,
            children,
        ) = folder_stack[-1]
        try:
            child = next(children)
        except StopIteration:
            folder_stack.pop()
            folder_count = len(visited_folders)
            if (
                progress_callback
                and folder_count % 100 == 0
                and folder_count != last_folder_progress
            ):
                progress_callback(
                    f"{item_count:,} itens lidos | "
                    f"{folder_count:,} pastas visitadas."
                )
                last_folder_progress = folder_count
            continue

        if cancel_callback and cancel_callback():
            raise ExportCancelled("Exportação cancelada pelo usuário.")
        child_name = child.get("name", "")
        child_mime = child.get("mimeType", "")
        child_path = f"{current_path}/{child_name}"
        child_segments = (*current_segments, child_name)
        size_bytes = parse_size(child.get("size"))
        shortcut = child.get("shortcutDetails", {})
        capabilities = child.get("capabilities", {})

        item_count += 1
        yield DriveItem(
            root_folder=source.root_name,
            source_type=source.source_type,
            drive_name=source.drive_name,
            parent_folder=current_name,
            path=child_path,
            depth=depth + 1,
            item_type=classify_item(child_mime),
            name=child_name,
            item_id=child.get("id", ""),
            mime_type=child_mime,
            size_bytes=size_bytes,
            size_readable=human_size(size_bytes),
            web_link=item_web_link(child),
            created_time=child.get("createdTime", ""),
            modified_time=child.get("modifiedTime", ""),
            owners=owners_to_text(child),
            drive_id=child.get("driveId", ""),
            parent_id=current_id,
            shortcut_target_id=shortcut.get("targetId", ""),
            shortcut_target_mime_type=shortcut.get("targetMimeType", ""),
            path_segments_json=json.dumps(
                child_segments, ensure_ascii=False, separators=(",", ":")
            ),
            resource_key=child.get("resourceKey", ""),
            shortcut_target_resource_key=shortcut.get("targetResourceKey", ""),
            can_download=capabilities.get("canDownload"),
            md5_checksum=child.get("md5Checksum", ""),
            sha256_checksum=child.get("sha256Checksum", ""),
            head_revision_id=child.get("headRevisionId", ""),
            version=str(child.get("version", "")),
        )

        child_id = child.get("id", "")
        if (
            child_mime == FOLDER_MIME_TYPE
            and child_id
            and child_id not in visited_folders
        ):
            visited_folders.add(child_id)
            folder_stack.append(
                (
                    child_id,
                    child_name,
                    child_path,
                    child_segments,
                    child.get("resourceKey", ""),
                    depth + 1,
                    iter(
                        list_children(
                            service,
                            child_id,
                            include_shared_drives,
                            drive_id=source.drive_id,
                            resource_key=child.get("resourceKey", ""),
                        )
                    ),
                )
            )

        if (
            progress_callback
            and progress_interval > 0
            and item_count % progress_interval == 0
        ):
            elapsed = max(time.monotonic() - started, 0.001)
            progress_callback(
                f"{item_count:,} itens lidos | "
                f"{len(visited_folders):,} pastas visitadas | "
                f"{item_count / elapsed:,.0f} itens/s"
            )

    if progress_callback:
        progress_callback(
            f"Leitura concluída: {item_count:,} itens em "
            f"{len(visited_folders):,} pastas."
        )


def build_drive_items(
    service,
    root_id: str,
    include_shared_drives: bool,
    progress_callback=None,
) -> list[DriveItem]:
    """Compatibilidade para integrações antigas; exportações usam streaming."""
    source = resolve_drive_source(service, root_id, include_shared_drives)
    return list(
        iter_drive_items(
            service,
            source,
            include_shared_drives,
            progress_callback=progress_callback,
        )
    )


EXCEL_HEADERS = [
    "Baixar?",
    "Formato de exportação",
    "Origem",
    "Drive compartilhado",
    "Pasta raiz",
    "Caminho completo",
    "Profundidade",
    "Pasta pai",
    "Tipo",
    "Nome",
    "Tamanho (bytes)",
    "Tamanho",
    "MIME type",
    "Link",
    "ID",
    "Criado em",
    "Modificado em",
    "Proprietário(s)",
    "Drive ID",
    "Parent ID",
    "Atalho: ID do alvo",
    "Atalho: MIME type do alvo",
    "Caminho (JSON)",
    "Resource key",
    "Atalho: resource key do alvo",
    "Pode baixar?",
    "SHA-256",
    "MD5",
    "Revisão",
    "Versão",
    "Schema",
]

EXCEL_COLUMN_WIDTHS = {
    "A": 12,
    "B": 22,
    "C": 24,
    "D": 30,
    "E": 24,
    "F": 60,
    "G": 14,
    "H": 28,
    "I": 20,
    "J": 42,
    "K": 18,
    "L": 14,
    "M": 38,
    "N": 18,
    "O": 38,
    "P": 22,
    "Q": 22,
    "R": 34,
    "S": 24,
    "T": 38,
    "U": 38,
    "V": 38,
}
EXCEL_LAST_COLUMN = "AE"
EXCEL_SCHEMA_VERSION = 2


def _plain_excel_cells(worksheet, values: Iterable[object]) -> list:
    """Cria células sem interpretar nomes iniciados por '=' como fórmulas."""
    from openpyxl.cell import WriteOnlyCell

    cells = []
    for value in values:
        if isinstance(value, str):
            value = EXCEL_ILLEGAL_CHARACTERS_RE.sub("", value)
        cell = WriteOnlyCell(worksheet, value=value)
        if isinstance(value, str):
            cell.data_type = "s"
        cells.append(cell)
    return cells


@lru_cache(maxsize=1)
def _inventory_row_styles():
    from openpyxl.styles import Alignment, Font, PatternFill

    return (
        PatternFill(fill_type="solid", fgColor="FFF2CC"),
        Font(bold=True, color="7F6000"),
        Alignment(horizontal="center"),
    )


@lru_cache(maxsize=16)
def _indent_alignment(depth: int):
    from openpyxl.styles import Alignment

    return Alignment(indent=min(max(depth, 0), 15))


def _create_inventory_sheet(workbook, index: int):
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    worksheet = workbook.create_sheet(f"Inventário {index}")
    worksheet.freeze_panes = "C2"
    worksheet.sheet_view.showGridLines = False
    for column, width in EXCEL_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width
    for column in ("W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"):
        worksheet.column_dimensions[column].hidden = True

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    selection_fill = PatternFill(fill_type="solid", fgColor="287D5A")
    technical_fill = PatternFill(fill_type="solid", fgColor="52606D")
    header_font = Font(color="FFFFFF", bold=True)
    header_cells = []
    for column_index, header in enumerate(EXCEL_HEADERS, start=1):
        cell = WriteOnlyCell(worksheet, value=header)
        if column_index <= 2:
            cell.fill = selection_fill
        elif column_index >= 23:
            cell.fill = technical_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        header_cells.append(cell)
    worksheet.append(header_cells)

    selection_validation = DataValidation(
        type="list", formula1='"Não,Sim"', allow_blank=True
    )
    export_validation = DataValidation(
        type="list",
        formula1='"AUTO,PDF,DOCX,XLSX,PPTX,PNG,ZIP,JSON,TXT,MP4"',
        allow_blank=True,
    )
    worksheet.data_validations.append(selection_validation)
    worksheet.data_validations.append(export_validation)
    selection_validation.add(f"A2:A{MAX_EXCEL_ROWS}")
    export_validation.add(f"B2:B{MAX_EXCEL_ROWS}")
    return worksheet


def _append_inventory_row(worksheet, row: DriveItem) -> None:
    from openpyxl.cell import WriteOnlyCell

    values: list[object] = [
        "Não",
        "AUTO",
        row.source_type,
        row.drive_name,
        row.root_folder,
        row.path,
        row.depth,
        row.parent_folder,
        row.item_type,
        row.name,
        row.size_bytes,
        row.size_readable,
        row.mime_type,
        "Abrir no Drive" if row.web_link else "",
        row.item_id,
        row.created_time,
        row.modified_time,
        row.owners,
        row.drive_id,
        row.parent_id,
        row.shortcut_target_id,
        row.shortcut_target_mime_type,
        row.path_segments_json,
        row.resource_key,
        row.shortcut_target_resource_key,
        row.can_download,
        row.sha256_checksum,
        row.md5_checksum,
        row.head_revision_id,
        row.version,
        EXCEL_SCHEMA_VERSION,
    ]
    cells = _plain_excel_cells(worksheet, values)
    selection_fill, selection_font, centered = _inventory_row_styles()
    cells[0].fill = selection_fill
    cells[0].font = selection_font
    cells[0].alignment = centered
    cells[1].alignment = centered
    if row.web_link:
        link_cell = WriteOnlyCell(worksheet, value="Abrir no Drive")
        link_cell.hyperlink = row.web_link
        link_cell.style = "Hyperlink"
        cells[13] = link_cell
    cells[9].alignment = _indent_alignment(min(max(row.depth, 0), 15))
    worksheet.append(cells)


def _write_instructions_sheet(worksheet) -> None:
    from openpyxl.styles import Font, PatternFill

    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 100
    rows = [
        ("DriveAutomate", "Como selecionar arquivos para download"),
        ("1", "Abra uma aba 'Inventário' e localize a coluna amarela 'Baixar?'."),
        ("2", "Marque como 'Sim' somente os arquivos desejados. O padrão seguro é 'Não'."),
        ("3", "Opcional: escolha um formato compatível para arquivos Google Workspace; AUTO usa o formato recomendado."),
        ("4", "Salve e feche o Excel antes de selecionar esta planilha no DriveAutomate."),
        ("Alternativa", "Também é possível apagar as linhas desnecessárias e, no aplicativo, escolher 'Todos os arquivos restantes'."),
        ("Importante", "Filtros visuais do Excel não removem linhas e não alteram a seleção."),
        ("Privacidade", "Este relatório contém nomes, IDs e possivelmente e-mails de proprietários. Guarde-o em local seguro."),
    ]
    for index, (label, value) in enumerate(rows):
        cells = _plain_excel_cells(worksheet, [label, value])
        if index == 0:
            for cell in cells:
                cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True, size=14)
        elif label in {"Importante", "Privacidade"}:
            cells[0].font = Font(bold=True, color="9C0006")
        worksheet.append(cells)


def _write_summary_sheet(
    worksheet,
    source: DriveSource | None,
    stats: InventoryStats,
    started_at: dt.datetime,
    finished_at: dt.datetime,
) -> None:
    source = source or DriveSource("", "", "", "", "", "")
    worksheet.sheet_view.showGridLines = False
    elapsed = finished_at - started_at
    summary_rows = [
        ("Relatório", "Inventário Google Drive — schema 2"),
        ("Origem detectada", source.source_type),
        ("Pasta raiz", source.root_name),
        ("ID da pasta raiz", source.root_id),
        ("Drive compartilhado", source.drive_name),
        ("Drive ID", source.drive_id),
        ("Proprietário da raiz", source.owner),
        ("Itens exportados", stats.item_count),
        ("Pastas", stats.folder_count),
        ("Arquivos/atalhos", stats.file_count),
        ("Soma dos tamanhos conhecidos", human_size(stats.total_size_bytes)),
        ("Abas de inventário", stats.worksheet_count),
        ("Início (UTC)", started_at.isoformat()),
        ("Fim (UTC)", finished_at.isoformat()),
        ("Duração", str(elapsed).split(".", maxsplit=1)[0]),
        (
            "Observação",
            "Arquivos Google Workspace podem não informar tamanho. A pasta raiz "
            "não é contada. Para baixar, marque 'Baixar?' como Sim.",
        ),
    ]
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 90
    for label, value in summary_rows:
        worksheet.append(
            _plain_excel_cells(worksheet, [label, value])
        )


def write_excel(
    rows: Iterable[DriveItem],
    output_path: str | Path,
    source: DriveSource | None = None,
    *,
    max_rows_per_sheet: int = MAX_EXCEL_ROWS,
) -> InventoryStats:
    """Grava o Excel em streaming e divide automaticamente relatórios gigantes."""
    from openpyxl import Workbook

    if max_rows_per_sheet <= EXCEL_HEADER_ROWS:
        raise ValueError("O limite de linhas por aba precisa permitir ao menos um item.")

    workbook = Workbook(write_only=True)
    instructions_sheet = workbook.create_sheet("Como usar")
    _write_instructions_sheet(instructions_sheet)
    summary_sheet = workbook.create_sheet("Resumo")
    worksheet = None
    rows_in_sheet = 0
    stats = InventoryStats()
    started_at = dt.datetime.now(tz=dt.timezone.utc)

    for row in rows:
        if worksheet is None or rows_in_sheet >= max_rows_per_sheet - EXCEL_HEADER_ROWS:
            if worksheet is not None:
                worksheet.auto_filter.ref = (
                    f"A1:{EXCEL_LAST_COLUMN}{rows_in_sheet + EXCEL_HEADER_ROWS}"
                )
            stats.worksheet_count += 1
            worksheet = _create_inventory_sheet(workbook, stats.worksheet_count)
            rows_in_sheet = 0

        _append_inventory_row(worksheet, row)
        rows_in_sheet += 1
        stats.item_count += 1
        stats.total_size_bytes += row.size_bytes or 0
        if row.mime_type == FOLDER_MIME_TYPE:
            stats.folder_count += 1
        else:
            stats.file_count += 1

    if worksheet is None:
        stats.worksheet_count = 1
        worksheet = _create_inventory_sheet(workbook, 1)
    worksheet.auto_filter.ref = (
        f"A1:{EXCEL_LAST_COLUMN}{rows_in_sheet + EXCEL_HEADER_ROWS}"
    )

    finished_at = dt.datetime.now(tz=dt.timezone.utc)
    _write_summary_sheet(summary_sheet, source, stats, started_at, finished_at)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return stats


def export_drive_inventory(
    folder_id_or_url: str,
    output_path: str | Path,
    include_shared_drives: bool = True,
    credentials_path: str | None = None,
    token_path: str | Path | None = None,
    client_config: dict[str, Any] | None = None,
    progress_callback=None,
    cancel_callback: Callable[[], bool] | None = None,
) -> int:
    """Fluxo completo usado pelo CLI e pela interface gráfica."""
    temporary_output: Path | None = None
    try:
        ensure_dependencies()

        resolved_credentials_path = credentials_path
        if resolved_credentials_path is None and client_config is None:
            resolved_credentials_path = str(oauth_client_path())
        creds = authenticate(
            credentials_path=resolved_credentials_path,
            token_path=token_path or default_token_path(),
            client_config=client_config,
        )
        service = build_drive_service(creds)
        source = resolve_drive_source(
            service,
            folder_id_or_url,
            include_shared_drives,
        )
        if progress_callback:
            location = source.source_type
            if source.drive_name:
                location += f" — {source.drive_name}"
            progress_callback(
                f"Origem detectada: {location}. Pasta: {source.root_name}"
            )

        rows = iter_drive_items(
            service,
            source,
            include_shared_drives,
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )
        output = Path(output_path)
        if output.suffix.lower() != ".xlsx":
            raise RuntimeError("O arquivo de saída precisa ter a extensão .xlsx.")
        output.parent.mkdir(parents=True, exist_ok=True)
        temporary_output = output.with_name(
            f".{output.stem}.parcial-{uuid.uuid4().hex}.xlsx"
        )
        stats = write_excel(rows, temporary_output, source=source)
        os.replace(temporary_output, output)
        temporary_output = None
        return stats.item_count
    except ExportCancelled:
        raise
    except (SystemExit, Exception) as exc:
        raise RuntimeError(format_user_error(exc)) from exc
    finally:
        if temporary_output and temporary_output.exists():
            try:
                temporary_output.unlink()
            except OSError:
                pass


def main() -> None:
    args = parse_args()
    started_at = dt.datetime.now(tz=dt.timezone.utc)

    exported_count = export_drive_inventory(
        folder_id_or_url=args.folder_id,
        output_path=args.output,
        include_shared_drives=args.include_shared_drives,
        credentials_path=args.credentials,
        token_path=args.token,
    )

    elapsed = dt.datetime.now(tz=dt.timezone.utc) - started_at
    print(f"Relatório criado: {os.path.abspath(args.output)}")
    print(f"Itens exportados: {exported_count}")
    print(f"Tempo de execução: {elapsed}")


if __name__ == "__main__":
    main()
