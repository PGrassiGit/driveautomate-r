import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from drive_inventory import (
    DriveItem,
    FOLDER_MIME_TYPE,
    build_drive_items,
    list_api_kwargs,
    normalize_folder_id,
    normalize_folder_reference,
    resolve_drive_source,
    write_excel,
)


class FakeRequest:
    def __init__(self, response):
        self.response = response
        self.headers = {}

    def execute(self, **_kwargs):
        return self.response


class FakeFiles:
    def __init__(self, metadata, children=None):
        self.metadata = metadata
        self.children = children or {}
        self.list_calls = []
        self.get_requests = []

    def get(self, fileId, **_kwargs):
        request = FakeRequest(self.metadata[fileId])
        self.get_requests.append(request)
        return request

    def list(self, **kwargs):
        self.list_calls.append(kwargs)
        query = kwargs["q"]
        parent_id = query.split("'")[1]
        return FakeRequest(
            {
                "files": self.children.get(parent_id, []),
                "incompleteSearch": False,
            }
        )


class FakeDrives:
    def __init__(self, names=None):
        self.names = names or {}

    def get(self, driveId, **_kwargs):
        return FakeRequest({"id": driveId, "name": self.names.get(driveId, "")})


class FakeService:
    def __init__(self, metadata, children=None, drive_names=None):
        self.fake_files = FakeFiles(metadata, children)
        self.fake_drives = FakeDrives(drive_names)

    def files(self):
        return self.fake_files

    def drives(self):
        return self.fake_drives


def folder_metadata(folder_id, name, **extra):
    return {
        "id": folder_id,
        "name": name,
        "mimeType": FOLDER_MIME_TYPE,
        "capabilities": {"canListChildren": True},
        **extra,
    }


def item(index, *, name=None, mime_type="application/pdf"):
    return DriveItem(
        root_folder="Raiz",
        source_type="Meu Drive",
        drive_name="",
        parent_folder="Raiz",
        path=f"Raiz/{name or f'Arquivo {index}'}",
        depth=1,
        item_type="Arquivo",
        name=name or f"Arquivo {index}",
        item_id=f"id-{index}",
        mime_type=mime_type,
        size_bytes=10,
        size_readable="10 B",
        web_link=f"https://drive.google.com/file/d/id-{index}/view",
        created_time="2026-01-01T00:00:00Z",
        modified_time="2026-01-02T00:00:00Z",
        owners="user@example.com",
        drive_id="",
        parent_id="root-id",
        shortcut_target_id="",
        shortcut_target_mime_type="",
    )


class DriveSourceTests(unittest.TestCase):
    def test_normalizes_example_style_urls(self):
        self.assertEqual(
            normalize_folder_id(
                "https://drive.google.com/drive/u/0/folders/example-folder-id"
            ),
            "example-folder-id",
        )

    def test_preserves_resource_key_from_shared_link(self):
        folder_id, resource_key = normalize_folder_reference(
            "https://drive.google.com/drive/folders/shared?resourcekey=example-key"
        )
        self.assertEqual((folder_id, resource_key), ("shared", "example-key"))

        service = FakeService(
            {"shared": folder_metadata("shared", "Compartilhada", ownedByMe=False)}
        )
        source = resolve_drive_source(
            service,
            "https://drive.google.com/drive/folders/shared?resourcekey=example-key",
        )
        self.assertEqual(source.root_resource_key, "example-key")
        self.assertEqual(
            service.fake_files.get_requests[0].headers["X-Goog-Drive-Resource-Keys"],
            "shared/example-key",
        )

    def test_classifies_owned_folder_as_my_drive(self):
        service = FakeService(
            {"personal": folder_metadata("personal", "Pessoal", ownedByMe=True)}
        )
        source = resolve_drive_source(service, "personal")
        self.assertEqual(source.source_type, "Meu Drive")
        self.assertEqual(source.drive_id, "")

    def test_classifies_shared_with_me_folder_without_drive_id(self):
        service = FakeService(
            {
                "shared": folder_metadata(
                    "shared",
                    "Compartilhada",
                    ownedByMe=False,
                    shared=True,
                    owners=[{"emailAddress": "owner@example.com"}],
                )
            }
        )
        source = resolve_drive_source(service, "shared")
        self.assertEqual(source.source_type, "Compartilhada comigo")
        self.assertEqual(source.owner, "owner@example.com")

    def test_classifies_shared_drive_and_uses_drive_corpus(self):
        service = FakeService(
            {
                "folder": folder_metadata(
                    "folder",
                    "Projetos",
                    driveId="shared-drive-id",
                    ownedByMe=False,
                )
            },
            children={
                "folder": [
                    {
                        "id": "file-1",
                        "name": "Documento.pdf",
                        "mimeType": "application/pdf",
                    }
                ]
            },
            drive_names={"shared-drive-id": "Drive da Equipe"},
        )

        rows = build_drive_items(service, "folder", include_shared_drives=True)

        self.assertEqual(rows[0].source_type, "Drive compartilhado")
        self.assertEqual(rows[0].drive_name, "Drive da Equipe")
        list_call = service.fake_files.list_calls[0]
        self.assertEqual(list_call["corpora"], "drive")
        self.assertEqual(list_call["driveId"], "shared-drive-id")
        self.assertTrue(list_call["includeItemsFromAllDrives"])
        self.assertTrue(list_call["supportsAllDrives"])

    def test_personal_or_shared_with_me_uses_user_corpus(self):
        kwargs = list_api_kwargs(True, drive_id="")
        self.assertEqual(kwargs["corpora"], "user")
        self.assertNotIn("driveId", kwargs)


class ExcelStreamingTests(unittest.TestCase):
    def test_streams_rows_splits_sheets_and_keeps_links(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "inventory.xlsx"
            stats = write_excel(
                [item(1), item(2), item(3, name="=NÃO_EXECUTAR()")],
                output,
                max_rows_per_sheet=3,
            )

            self.assertEqual(stats.item_count, 3)
            self.assertEqual(stats.worksheet_count, 2)

            workbook = load_workbook(output, read_only=False, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["Como usar", "Resumo", "Inventário 1", "Inventário 2"],
            )
            first_sheet = workbook["Inventário 1"]
            self.assertEqual(first_sheet["A2"].value, "Não")
            self.assertEqual(first_sheet["B2"].value, "AUTO")
            self.assertEqual(first_sheet["N2"].value, "Abrir no Drive")
            self.assertEqual(
                first_sheet["N2"].hyperlink.target,
                "https://drive.google.com/file/d/id-1/view",
            )
            self.assertTrue(first_sheet.column_dimensions["W"].hidden)
            self.assertEqual(first_sheet.freeze_panes, "C2")
            validation_formulas = {
                validation.formula1 for validation in first_sheet.data_validations.dataValidation
            }
            self.assertIn('"Não,Sim"', validation_formulas)
            second_sheet = workbook["Inventário 2"]
            self.assertEqual(second_sheet["J2"].value, "=NÃO_EXECUTAR()")
            self.assertEqual(second_sheet["J2"].data_type, "s")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
